""" Simulations where at least one enzyme is randomly distributed

:Author: Yin Hoon Chew <yinhoon.chew@bham.ac.uk>
:Date: 07-11-2022
:License: MIT

"""

from model import *
from openpyxl import Workbook
from os.path import dirname, abspath, join
from sklearn.preprocessing import PolynomialFeatures, StandardScaler
import collections
import itertools
import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
import seaborn as sns
import statsmodels.api as sm
import sys

# Find code directory relative to our directory
THIS_DIR = dirname(__file__)
CODE_DIR = abspath(join(THIS_DIR, '..', 'discretised_fba'))
sys.path.append(CODE_DIR)
from discretised_fba import DiscretisedCell

POPULATION_SIZE = 100
ENZYME_DISTRIBUTION = ['uniform', 'steep_inside_out', 'steep_outside_in', 'random']
DIFFUSING_METABOLITES = [['A[c]', 'B[c]'], ['A[c]'], ['B[c]'], []]  

def simulation():
    """ Run simulation

    Returns:
        :obj:`dict`: simulation results
    """
    # initialise results dictionary
    permutations = list(itertools.product(ENZYME_DISTRIBUTION, repeat=3))

    # remove gradient distribution for transport protein        
    results = {','.join(i): {' and '.join(j): [] for j in DIFFUSING_METABOLITES} \
        for i in permutations if i[0] not in ['steep_inside_out', 'steep_outside_in'] \
        and 'random' in i}

    # Run simulations
    for diffusion in DIFFUSING_METABOLITES:
        diff_name = ' and '.join(diffusion)
        for shape in CELL_SHAPES:
            cell = DiscretisedCell(shape['ShapeID'], shape['Width'], shape['Length'])
            cell.create_reactions(METABOLITES[0], METABOLITES[1:], EXCHANGE_REACTION,
                INTRACELLULAR_REACTIONS, 'Biomass_reaction')
            cell.create_transport_reactions(['A[c]'], TRANSPORT_REACTION)
            cell.create_diffusion(diffusion)        
            
            for distrib_name in results:
                distribution_combination = distrib_name.split(',')
                all_solutions = []
                
                for individual in range(POPULATION_SIZE):                       
                    for ind, protein in enumerate(CELLULAR_PROTEIN_CONCENTRATION):
                        distribution = distribution_combination[ind]
                        all_regions = False if ind == 0 else True # transport protein
                        random_distribution = True if distribution == 'random' else False
                        if distribution == 'moderate_inside_out':
                            gradient = -0.5
                        elif distribution == 'steep_inside_out':
                            gradient = -1.0
                        elif distribution == 'moderate_outside_in':
                            gradient = 0.5
                        elif distribution == 'steep_outside_in':
                            gradient = 1.0
                        else:
                            gradient = 0.0    
                        cell.distribute_enzyme(protein[0], protein[1], gradient=gradient, 
                            random_distribution=random_distribution, all_regions=all_regions)
                    synth_rxns, trans_rxns = cell.calc_enzymatic_bounds(PROTEIN_REACTION_RULE)
                    rxn_bounds = {}
                    for rxn in cell.model.reactions:
                        if rxn.id in synth_rxns:
                            rxn_bounds[rxn.id] = (0, synth_rxns[rxn.id])
                        elif rxn.id in trans_rxns:
                            rxn_bounds[rxn.id] = (0, trans_rxns[rxn.id])
                        else:
                            rxn_bounds[rxn.id] = (0, 1000.)
                    cell.set_bounds(rxn_bounds)
                    all_solutions.append(cell.solve())
                results[distrib_name][diff_name].append({
                    'ShapeID': shape['ShapeID'],
                    'Aspect ratio': cell.aspect_ratio,
                    'Perimeter-to-area ratio': cell.peri_to_area,
                    'Simulation': all_solutions,
                    })

    return results            
            

def write_results(results, model_name):
    """ Write results to an excel file

    Args:
        results (:obj:`dict`): simulation results
        model_name (:obj:`str`): name of model
    """
    dest_filename = abspath(join(THIS_DIR, '..', 'results', 
        f'{model_name}_stochastic_steep_gradients.xlsx'))
    wb = Workbook()
    sheets = {}

    column_index = {distrib_name: ind+1  for ind, distrib_name in enumerate(results)}
    for diffusion in DIFFUSING_METABOLITES:
        diff_name = ' and '.join([i[0:i.index('[')] for i in diffusion])
        diff_name = diff_name if diff_name else 'No diffusion'
        for shape in CELL_SHAPES:
            sheet_name = f"{shape['ShapeID']}({diff_name})"
            ws = wb.create_sheet(title=sheet_name)    
            sheets[sheet_name] = ws
            for distrib_name, ind in column_index.items():
                _ = ws.cell(column=ind, row=1, 
                            value=distrib_name)

    for distrib_name, data in results.items():
        col = column_index[distrib_name]
        for diff, details in data.items():
            diffusion_list = diff.split(' and ')
            diff_name = ' and '.join([i[0:i.index('[')] for i in diffusion_list if i])
            diff_name = diff_name if diff_name else 'No diffusion'
            for shape in details:
                sheet_name = f"{shape['ShapeID']}({diff_name})"
                ws = sheets[sheet_name]
                row = 2
                for individual in shape['Simulation']:
                    _ = ws.cell(column=col, row=row, 
                            value=individual.objective_value)
                    row += 1        
            
    wb.save(filename = dest_filename)


def plot_kernel_density(results, model_name):
    """ Plot kernel density for each cell geometry

    Args:
        results (:obj:`dict`): simulation results
        model_name (:obj:`str`): name of model
    """
    distribution_dictionary = {
            'uniform': '0', 
            'steep_inside_out': '\u2013', 
            'steep_outside_in': '+',
            'random': 'r'}
    cols = results.keys()
    colours = ['r', 'g', 'b', 'k']

    for shape in CELL_SHAPES:
        fig, axes = plt.subplots(nrows=4, ncols=6, figsize=(10, 6), sharey=True)
        axes = axes.ravel()  # array to 1D
        fig.suptitle(f'Cell geometry: {shape["ShapeID"]}')
        for col, ax in zip(cols, axes):
            for diff, colour in zip(sorted(results[col].keys()), colours):
                value = results[col][diff]
                diff_name = diff if diff else 'No diffusion'
                data = [j.objective_value for i in value for j in i['Simulation'] \
                    if i['ShapeID']==shape['ShapeID']]
                if np.var(data) < 1e-05:
                    ax.axvline(np.mean(data), color=colour, label=diff_name)
                else:
                    sns.set_style('white')
                    sns.kdeplot(data, ax=ax, color=colour, label=diff_name)
            subplot_title = ''.join([distribution_dictionary[i] for i in col.split(',')])
            ax.title.set_text(subplot_title)
            ax.set_ylim([0, 0.3])
            ax.set_xlim([0, 120])
            line, label = ax.get_legend_handles_labels()
        fig.legend(line, label, loc='lower right', bbox_to_anchor=(1.0, 0.))
        fig.tight_layout()
        axes.flat[-1].set_visible(False)
        dest_figname = abspath(join(THIS_DIR, '..', 'results', 
            f'{model_name}_{shape["ShapeID"]}.tiff'))
        fig.savefig(dest_figname, dpi=1200)
    

def plot_cv(results, model_name):
    """Plot coefficient of variation

    Args:
        results (:obj:`dict`): simulation results
        model_name (:obj:`str`): name of model
    """
    distribution_dictionary = {
            'uniform': '0', 
            'steep_inside_out': '\u2013', 
            'steep_outside_in': '+',
            'random': 'r'}
    cols = results.keys()
    colours = ['r', 'g', 'b', 'k']
    fig, axes = plt.subplots(nrows=4, ncols=6, figsize=(10, 6), sharey=True)
    axes = axes.ravel()  # array to 1D
    fig.suptitle(model_name)
    fig.supylabel('Coefficient of variation')
    fig.supxlabel('Perimeter-to-area ratio')
    cv = lambda x: np.std(x, ddof=1) / np.mean(x) * 100

    for col, ax in zip(cols, axes):
        for diff, colour in zip(sorted(results[col].keys()), colours):
            value = results[col][diff]
            diff_name = diff if diff else 'No diffusion'
            ax.plot([i['Perimeter-to-area ratio'] for i in value], 
                [cv([j.objective_value for j in i['Simulation']]) for i in value], 
                colour, label=diff_name)
        subplot_title = ''.join([distribution_dictionary[i] for i in col.split(',')])
        ax.title.set_text(subplot_title)
        ax.set_ylim([0, 15])    
        line, label = ax.get_legend_handles_labels()
    fig.legend(line, label, loc='lower right', bbox_to_anchor=(1.0, 0.))
    fig.tight_layout()
    axes.flat[-1].set_visible(False)
    dest_figname = abspath(join(THIS_DIR, '..', 'results', f'{model_name}_cv.tiff'))
    fig.savefig(dest_figname, dpi=1200)


def plot_std(results, model_name):
    """Plot standard deviation

    Args:
        results (:obj:`dict`): simulation results
        model_name (:obj:`str`): name of model
    """
    distribution_dictionary = {
            'uniform': '0', 
            'steep_inside_out': '\u2013', 
            'steep_outside_in': '+',
            'random': 'r'}
    cols = results.keys()
    colours = ['r', 'g', 'b', 'k']
    fig, axes = plt.subplots(nrows=4, ncols=6, figsize=(10, 6), sharey=True)
    axes = axes.ravel()  # array to 1D
    fig.suptitle(model_name)
    fig.supylabel('Standard deviation')
    fig.supxlabel('Perimeter-to-area ratio')

    for col, ax in zip(cols, axes):
        for diff, colour in zip(sorted(results[col].keys()), colours):
            value = results[col][diff]
            diff_name = diff if diff else 'No diffusion'
            ax.plot([i['Perimeter-to-area ratio'] for i in value], 
                [np.std([j.objective_value for j in i['Simulation']], ddof=1) for i in value], 
                colour, label=diff_name)
        subplot_title = ''.join([distribution_dictionary[i] for i in col.split(',')])
        ax.title.set_text(subplot_title)
        ax.set_ylim([0, 8])    
        line, label = ax.get_legend_handles_labels()
    fig.legend(line, label, loc='lower right', bbox_to_anchor=(1.0, 0.))
    fig.tight_layout()
    axes.flat[-1].set_visible(False)
    dest_figname = abspath(join(THIS_DIR, '..', 'results', f'{model_name}_std.tiff'))
    fig.savefig(dest_figname, dpi=1200)


if __name__ == '__main__':
    # simulate default model
    model_name = 'default_model'
    results = simulation()
    write_results(results, model_name)
    plot_kernel_density(results, model_name)
    plot_cv(results, model_name)
    plot_std(results, model_name)

    # simulate model with metabolite A as a limiting factor
    INTRACELLULAR_REACTIONS = {
    'Synthesise_B': '2 A[c] --> B[c]',
    'Synthesise_biomass': 'B[c] --> Biomass',
    'Biomass_reaction': 'Biomass --> '
    }
    model_name = 'A_limiting_model'
    results = simulation()
    plot_cv(results, model_name)
    plot_std(results, model_name)

    # simulate model with metabolite A as a limiting factor
    INTRACELLULAR_REACTIONS = {
    'Synthesise_B': 'A[c] --> B[c]',
    'Synthesise_biomass': '2 B[c] --> Biomass',
    'Biomass_reaction': 'Biomass --> '
    }
    model_name = 'B_limiting_model'
    results = simulation()
    plot_cv(results, model_name)
    plot_std(results, model_name)
