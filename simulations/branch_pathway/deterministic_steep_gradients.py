""" Simulations where enzymes are distributed into
uniform or steep gradient patterns

:Author: Yin Hoon Chew <yinhoon.chew@bham.ac.uk>
:Date: 07-11-2022
:License: MIT

"""

from model import *
from openpyxl import Workbook
from os.path import dirname, abspath, join
import collections
import itertools
import matplotlib.pyplot as plt
import pkg_resources
import sys

# Find code directory relative to our directory
THIS_DIR = dirname(__file__)
CODE_DIR = abspath(join(THIS_DIR, '..', '..', 'discretised_fba'))
sys.path.append(CODE_DIR)
from discretised_fba import DiscretisedCell


ENZYME_DISTRIBUTION = ['uniform', 'steep_inside_out', 'steep_outside_in']
DIFFUSING_METABOLITES = [['X[c]', 'Y[c]'], ['X[c]'], ['Y[c]'], []]  

# initialise results dictionary
permutations = list(itertools.product(ENZYME_DISTRIBUTION, repeat=4))

# remove gradient distribution for transport protein        
results = {','.join(i): {' and '.join(j): [] for j in DIFFUSING_METABOLITES} \
    for i in permutations if i[0] not in ['steep_inside_out', 'steep_outside_in']} 

# Run simulations
for diffusion in DIFFUSING_METABOLITES:
    diff_name = ' and '.join(diffusion)
    for shape in CELL_SHAPES:
        cell = DiscretisedCell(shape['ShapeID'], shape['Width'], shape['Length'])
        cell.create_reactions(METABOLITES[0], METABOLITES[1:], EXCHANGE_REACTION,
            INTRACELLULAR_REACTIONS, {'Demand_energy': 1.0, 'Demand_toxin': -2.0})
        cell.create_transport_reactions(['X[c]'], TRANSPORT_REACTION)
        cell.create_diffusion(diffusion)        
        
        for distrib_name in results:
            distribution_combination = distrib_name.split(',')            
            for ind, protein in enumerate(CELLULAR_PROTEIN_CONCENTRATION):
                distribution = distribution_combination[ind]
                all_regions = False if ind == 0 else True # transport protein
                random_distribution = True if distribution == 'random' else False
                if distribution == 'moderate_inside_out':
                    gradient = -0.5
                elif distribution == 'steep_inside_out':
                    gradient = -1.0
                elif distribution == 'moderate_outside_in':
                    gradient = 0.5
                elif distribution == 'steep_outside_in':
                    gradient = 1.0
                else:
                    gradient = 0.0    
                cell.distribute_enzyme(protein[0], protein[1], gradient=gradient, 
                    random_distribution=random_distribution, all_regions=all_regions)
            synth_rxns, trans_rxns = cell.calc_enzymatic_bounds(PROTEIN_REACTION_RULE)
            rxn_bounds = {}
            for rxn in cell.model.reactions:
                if rxn.id in synth_rxns:
                    rxn_bounds[rxn.id] = (0, synth_rxns[rxn.id])
                elif rxn.id in trans_rxns:
                    rxn_bounds[rxn.id] = (0, trans_rxns[rxn.id])
                else:
                    rxn_bounds[rxn.id] = (0, 1000.)
            cell.set_bounds(rxn_bounds)
            solution = cell.solve()
            results[distrib_name][diff_name].append({
                'ShapeID': shape['ShapeID'],
                'Aspect ratio': cell.aspect_ratio,
                'Perimeter-to-area ratio': cell.peri_to_area,
                'Simulation': solution,
                })
            
# Write results to an excel file
dest_filename = abspath(join(THIS_DIR, '..', '..', 'results', 'branch_pathway', 
    'deterministic_steep_gradients.xlsx'))
wb = Workbook()
sheets = {}
shape_properties = ['ShapeID', 'Aspect ratio', 'Perimeter-to-area ratio']
for diffusion in DIFFUSING_METABOLITES:
    diff_name = ' and '.join(diffusion)
    sheet_name = ' and '.join([i[0:i.index('[')] for i in diffusion])
    ws = wb.create_sheet(title=sheet_name if sheet_name else 'No diffusion')    
    sheets[diff_name] = ws    
    for ind, val in enumerate(shape_properties):
        ws[f'A{ind+1}'] = val 

row = 4
data_entry = ['Objective value', 'Primary flux', 'Secondary flux']
for distrib_name, data in results.items():
    for diff, details in data.items():
        sheets[diff][f'A{row}'] = distrib_name
        for ind, val in enumerate(data_entry):
            sheets[diff][f'B{row+ind}'] = val
        for shape_no, shape in enumerate(details):
            for ind, val in enumerate(shape_properties):
                _ = sheets[diff].cell(
                        column=shape_no+3, row=ind+1, 
                        value=shape[val])
            # Write objective value 
            _ = sheets[diff].cell(
                    column=shape_no+3, row=row, 
                    value=shape['Simulation'].objective_value)
            # Write flux through primary energy production path 
            _ = sheets[diff].cell(
                    column=shape_no+3, row=row+1, 
                    value=sum([shape['Simulation'].fluxes[i] \
                        for i in shape['Simulation'].fluxes.keys() \
                        if 'Primary' in i]))
            # Write flux through secondary energy production path
            _ = sheets[diff].cell(
                    column=shape_no+3, row=row+2, 
                    value=sum([shape['Simulation'].fluxes[i] \
                        for i in shape['Simulation'].fluxes.keys() \
                        if 'Secondary' in i]))
    row += 3       
        
wb.save(filename = dest_filename)

# Plot results
distribution_dictionary = {
        'uniform': '0', 
        'steep_inside_out': '\u2013', 
        'steep_outside_in': '+'}
colours = ['crimson', 'seagreen', 'cornflowerblue', 'orange']
styles = ['dashed', 'solid', 'dashed', 'solid']  
cols = results.keys()      

# line graph for total energy produced
fig, axes = plt.subplots(nrows=4, ncols=7, figsize=(10, 6), sharey=True)
axes = axes.ravel()  # array to 1D
fig.supylabel('Total energy produced')
fig.supxlabel('Perimeter-to-area ratio')

for col, ax in zip(cols, axes):
    for diff, colour, style in zip(sorted(results[col].keys()), colours, styles):
        value = results[col][diff]
        diff_name = diff if diff else 'No diffusion'
        
        total_energy = []
        for i in value:
            total_primary_flux = sum([i['Simulation'].fluxes[j] \
                for j in i['Simulation'].fluxes.keys() \
                if 'Primary' in j])
            total_secondary_flux = sum([i['Simulation'].fluxes[j] \
                for j in i['Simulation'].fluxes.keys() \
                if 'Secondary' in j])
            total_energy.append(4*total_primary_flux + 3*total_secondary_flux)
        
        ax.plot([i['Perimeter-to-area ratio'] for i in value], 
            total_energy, colour, label=diff_name, 
            linestyle=style, linewidth=3, alpha=0.35)
    subplot_title = ''.join([distribution_dictionary[i] for i in col.split(',')])
    ax.set_ylim([100, 450])  
    ax.title.set_text(subplot_title)
    line, label = ax.get_legend_handles_labels()
fig.legend(line, label, loc='lower right', bbox_to_anchor=(1.0, 0.))
fig.tight_layout()
axes.flat[-1].set_visible(False)
dest_figname = abspath(join(THIS_DIR, '..', '..', 'results', 'branch_pathway', 
    'deterministic_steep_gradients_energy.png'))
fig.savefig(dest_figname, dpi=1200)

# line graph for flux ratio
fig, axes = plt.subplots(nrows=4, ncols=7, figsize=(10, 6), sharey=True)
axes = axes.ravel()  # array to 1D
fig.supylabel('Secondary-to-primary flux ratio')
fig.supxlabel('Perimeter-to-area ratio')

for col, ax in zip(cols, axes):
    for diff, colour, style in zip(sorted(results[col].keys()), colours, styles):
        value = results[col][diff]
        diff_name = diff if diff else 'No diffusion'
        
        flux_ratio = []
        for i in value:
            total_primary_flux = sum([i['Simulation'].fluxes[j] \
                for j in i['Simulation'].fluxes.keys() \
                if 'Primary' in j])
            total_secondary_flux = sum([i['Simulation'].fluxes[j] \
                for j in i['Simulation'].fluxes.keys() \
                if 'Secondary' in j])
            flux_ratio.append(total_secondary_flux/total_primary_flux)
        
        ax.plot([i['Perimeter-to-area ratio'] for i in value], 
            flux_ratio, colour, label=diff_name, 
            linestyle=style, linewidth=3, alpha=0.35)
    subplot_title = ''.join([distribution_dictionary[i] for i in col.split(',')])
    ax.set_ylim([0, 1.0])  
    ax.title.set_text(subplot_title)
    line, label = ax.get_legend_handles_labels()
fig.legend(line, label, loc='lower right', bbox_to_anchor=(1.0, 0.))
fig.tight_layout()
axes.flat[-1].set_visible(False)
dest_figname = abspath(join(THIS_DIR, '..', '..', 'results', 'branch_pathway', 
    'deterministic_steep_gradients_ratio.png'))
fig.savefig(dest_figname, dpi=1200)

# Bar graph
fig, axes = plt.subplots(nrows=1, ncols=3, figsize=(10, 5))
pad = 5 # in points
cols = ['X[c]', 'Y[c]', 'No diffusion']
rows = ['6x6']

for row, shape in enumerate(rows):
    for col, diff_name in enumerate(cols):
        diff_key = diff_name if diff_name != 'No diffusion' else ''
        data1 = collections.defaultdict(list)
        data2 = collections.defaultdict(list)
        for distrib_name, values in results.items():
            for cell in values[diff_key]:
                if cell['ShapeID'] == shape:
                    total_primary_flux = sum([cell['Simulation'].fluxes[i] \
                        for i in cell['Simulation'].fluxes.keys() \
                        if 'Primary' in i])
                    total_secondary_flux = sum([cell['Simulation'].fluxes[i] \
                        for i in cell['Simulation'].fluxes.keys() \
                        if 'Secondary' in i])
                    total_energy = 4*total_primary_flux + 3*total_secondary_flux
                    flux_ratio = total_secondary_flux/total_primary_flux 
                    data1[round(total_energy, 5)].append(
                        ''.join([distribution_dictionary[i] 
                            for i in distrib_name.split(',')]))
                    data2[round(total_energy, 5)].append(flux_ratio)
        y1 = sorted(data1.keys())
        y2 = [sum(data2[i])/len(data2[i]) for i in y1]
        x = ['\n'.join(data1[i]) for i in y1]
        ax = axes[col]
        ax2 = ax.twinx()
        ax.bar(x, y1, width=-0.4, color='cornflowerblue', alpha=0.35, align='edge')
        ax2.bar(x, y2, width=0.4, color='crimson', alpha=0.35, align='edge')
        ax.set_ylim([0, 450])
        ax2.set_ylim([0, 1.0])
        ax.tick_params(axis='x', labelsize=7)
        ax.title.set_text(diff_name)

fig.supylabel(x=0.03, y=0.6, t='Total energy produced', size=11)
fig.text(x=0.94, y=0.6, s='Secondary-to-primary flux ratio\n\n', size=11, rotation=270, ha='center', va='center')
plt.subplots_adjust(left=0.1,
                    bottom=0.3,
                    right=0.9,
                    top=0.9,
                    wspace=0.4,
                    hspace=0.4)

dest_figname = abspath(join(THIS_DIR, '..', '..', 'results', 'branch_pathway', 
    'deterministic_steep_gradients_bar.png'))
fig.savefig(dest_figname, dpi=1200)

plt.show()
