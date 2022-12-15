""" Design of simulation experiments

:Author: Yin Hoon Chew <yinhoon.chew@bham.ac.uk>
:Date: 07-11-2022
:License: MIT

"""
from os.path import dirname, abspath, join
import sys

# Find code directory relative to our directory
THIS_DIR = dirname(__file__)
CODE_DIR = abspath(join(THIS_DIR, '..', 'discretised_fba'))
sys.path.append(CODE_DIR)


from discretised_fba import DiscretisedCell
from openpyxl import Workbook
import itertools
import matplotlib.pyplot as plt


# Exoerimental designs
METABOLITES = ['A[e]', 'A[c]', 'B[c]', 'Biomass']
EXCHANGE_REACTION = {'EX_A': ' --> A[e]'}
TRANSPORT_REACTION = {'Transport_A': 'A[e] --> A[c]'}
INTRACELLULAR_REACTIONS = {
    'Synthesise_B': 'A[c] --> B[c]',
    'Synthesise_biomass': 'B[c] --> Biomass',
    'Biomass_reaction': 'Biomass --> '
    }
PROTEIN_REACTION_RULE = {
    'Transport_A': {'Protein1': 1.0},
    'Synthesise_B': {'Protein2': 1.0},
    'Synthesise_biomass': {'Protein3': 1.0}
    }
CELLULAR_PROTEIN_CONCENTRATION = [
    ('Protein1', 100),
    ('Protein2', 100), 
    ('Protein3', 100),
]
CELL_SHAPES = [
    {'ShapeID': '6x6', 'Width': 6, 'Length': 6},
    {'ShapeID': '4x9', 'Width': 4, 'Length': 9},
    {'ShapeID': '3x12', 'Width': 3, 'Length': 12},
    {'ShapeID': '2x18', 'Width': 2, 'Length': 18},
    {'ShapeID': '1x36', 'Width': 1, 'Length': 36},
]
#ENZYME_DISTRIBUTION = ['uniform', 'moderate_inside_out', 'steep_inside_out', 
#    'moderate_outside_in','steep_outside_in', 'random']
ENZYME_DISTRIBUTION = ['uniform', 'steep_inside_out', 'steep_outside_in']
DIFFUSING_METABOLITES = [['A[c]', 'B[c]'], ['A[c]'], ['B[c]'], []]  

# initialise results dictionary
permutations = list(itertools.product(ENZYME_DISTRIBUTION, repeat=3))

# remove gradient distribution for transport protein        
results = {','.join(i): {' and '.join(j): [] for j in DIFFUSING_METABOLITES} \
    for i in permutations if i[0] not in ['steep_inside_out', 'steep_outside_in']} 

# Run simulations
for diffusion in DIFFUSING_METABOLITES:
    diff_name = ' and '.join(diffusion)
    for shape in CELL_SHAPES:
        cell = DiscretisedCell(shape['ShapeID'], shape['Width'], shape['Length'])
        cell.create_reactions(METABOLITES[0], METABOLITES[1:], EXCHANGE_REACTION,
            INTRACELLULAR_REACTIONS, 'Biomass_reaction')
        cell.create_transport_reactions(['A[c]'], TRANSPORT_REACTION)
        cell.create_diffusion(diffusion)        
        
        for distrib_name in results:
            distribution_combination = distrib_name.split(',')            
            for ind, protein in enumerate(CELLULAR_PROTEIN_CONCENTRATION):
                distribution = distribution_combination[ind]
                all_regions = False if ind == 0 else True # transport protein
                random_distribution = True if distribution == 'random' else False
                if distribution == 'moderate_inside_out':
                    gradient = -0.5
                elif distribution == 'steep_inside_out':
                    gradient = -1.0
                elif distribution == 'moderate_outside_in':
                    gradient = 0.5
                elif distribution == 'steep_outside_in':
                    gradient = 1.0
                else:
                    gradient = 0.0    
                cell.distribute_enzyme(protein[0], protein[1], gradient=gradient, 
                    random_distribution=random_distribution, all_regions=all_regions)
            synth_rxns, trans_rxns = cell.calc_enzymatic_bounds(PROTEIN_REACTION_RULE)
            rxn_bounds = {}
            for rxn in cell.model.reactions:
                if rxn.id in synth_rxns:
                    rxn_bounds[rxn.id] = (0, synth_rxns[rxn.id])
                elif rxn.id in trans_rxns:
                    rxn_bounds[rxn.id] = (0, trans_rxns[rxn.id])
                else:
                    rxn_bounds[rxn.id] = (0, 1000.)
            cell.set_bounds(rxn_bounds)
            solution = cell.solve()
            results[distrib_name][diff_name].append({
                'ShapeID': shape['ShapeID'],
                'Aspect ratio': cell.aspect_ratio,
                'Perimeter-to-area ratio': cell.peri_to_area,
                'Simulation': solution,
                })
            
# Write results to an excel file
dest_filename = 'results.xlsx'
wb = Workbook()
sheets = {}
shape_properties = ['ShapeID', 'Aspect ratio', 'Perimeter-to-area ratio']
for diffusion in DIFFUSING_METABOLITES:
    diff_name = ' and '.join(diffusion)
    sheet_name = ' and '.join([i[0:i.index('[')] for i in diffusion])
    ws = wb.create_sheet(title=sheet_name if sheet_name else 'No diffusion')    
    sheets[diff_name] = ws    
    for ind, val in enumerate(shape_properties):
        ws[f'A{ind+1}'] = val 

row = 4
for distrib_name, data in results.items():
    for diff, details in data.items():
        sheets[diff][f'A{row}'] = distrib_name
        for shape_no, shape in enumerate(details):
            for ind, val in enumerate(shape_properties):
                _ = sheets[diff].cell(
                        column=shape_no+2, row=ind+1, 
                        value=shape[val])
            _ = sheets[diff].cell(
                    column=shape_no+2, row=row, 
                    value=shape['Simulation'].objective_value)
    row += 1        
        
wb.save(filename = dest_filename)

# Plot results
fig, axes = plt.subplots(nrows=3, ncols=3, figsize=(10, 6))
axes = axes.ravel()  # array to 1D
cols = results.keys()
colours = ['r', 'g', 'b', 'k']

for col, ax in zip(cols, axes):
    for diff, colour in zip(sorted(results[col].keys()), colours):
        value = results[col][diff]
        diff_name = diff if diff else 'No diffusion'
        ax.plot([i['Perimeter-to-area ratio'] for i in value], 
            [i['Simulation'].objective_value for i in value], colour, label=diff_name)
    distribution_dictionary = {
        'uniform': 'u', 
        'steep_inside_out': 'io', 
        'steep_outside_in': 'oi'}
    subplot_title = ','.join([distribution_dictionary[i] for i in col.split(',')])    
    ax.title.set_text(subplot_title)
    ax.set_xlabel('Perimeter-to-area ratio')
    ax.set_ylabel('Biomass growth')
ax.legend(bbox_to_anchor=(1.05, 0), loc='lower left', borderaxespad=0.)
fig.tight_layout()
plt.show()
