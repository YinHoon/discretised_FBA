""" Methods for running discretised FBA

:Author: Yin Hoon Chew <yinhoon.chew@bham.ac.uk>
:Author: Alexandra Anamaria-Sorinca
:Date: 13-09-2022
:License: MIT

"""

import cobra
import re
import numpy as np
from numpy import matrix
from numpy import linalg
import openpyxl
cobra_config = cobra.Configuration()
from cobra import Model, Reaction, Metabolite
from cobra.util.solver import linear_reaction_coefficients
from cobra.io import load_model
from cobra.flux_analysis import gapfill
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
model = load_model("textbook")


#create the shape of the cell in a matrix form
#the row size * the column size = the number of cell compartiments
def run_model(r_size, c_size, enzyme):
    r_size = r_size
    c_size = c_size

    '''
    #if ij=1 then that part of a cell exists, otherwise ij=0
    #if ij=0 set the fluxes to be 0
    
    d = np.ones((int(r_size),int(c_size)))
    d[0,1] = '0'
    print(d)
    '''

    #define the reactions and their fluxes
    #set fluxes upper/lower bounds
    #metabolites create within the reaction models


    test_model = cobra.Model("test_model")

    # create empty dictionary for metabolite
    met = {}
    # create empty dictionary for reactions
    react = {}
    # create Ae and Re separately and add them to the met and react dict
    # Ae and Re are separated because they are from the external environment, which for simplicity is set to be neutral

    Ae = cobra.Metabolite('Ae')
    test_model.add_metabolites([Ae])
    met[Ae.id]=Ae
    rE = cobra.Reaction('rE')
    test_model.add_reactions([rE])
    rE.reaction = " --> Ae"
    react[rE.id]=rE

    #create names for the rest of the reactions, which are their id's
    #the react object in the dictionary is the reaction itself, defined by the metabolites
    #cobrapy has a tool which enables us to define the metabolites from the reaction....
    #instead of defining the metabolites manually

    for i in range(r_size):
        for j in range(c_size):
            #create name for metA ,B,C for compartment ij
            mets = ['A[c]','B[c]','Biomass[c]']
            for m in mets:
                met_id = m + str (i+1) + str (j+1)
                met_obj = cobra.Metabolite(met_id)
                test_model.add_metabolites([met_obj])
                met[met_id] = met_obj
            r_name = {'rAeAc': "Ae --> A[c]",
                      'rAcBc': "A[c] --> B[c]",
                      'rBcBio': "2 B[c] --> Biomass[c]",
                      'rBioE': "Biomass[c] -->  "}
            for k,v in r_name.items():
                react_id = k + str (i+1) + str(j+1)
                react_obj = cobra.Reaction(react_id)
                test_model.add_reactions([react_obj])
                react[react_id] = react_obj
                result = re.sub("]", "]" + str(i+1) + str(j+1), v)
                react_obj.reaction = result
                print(result)
    print(react)

    #enzymes can be distributed non-uniformly across the cell....
    #....which causes the geometric non-uniformity of the cell
    # in the excel workbook (appendix), there are a few ways to distribute the enzymes
    # this is currently done manually, each value is being inserted in the terminal after running the code

    enzyme_concentration = {'k_AeAc': 1000, 'k_AcBc':2000}
    distrib = {}
    for m in enzyme_concentration:
        for i in range(r_size):
            for j in range(c_size):
                distrib_id = m + str(i+1) + str(j+1)
                distrib[distrib_id]=enzyme[distrib_id]
        if not np.testing.assert_almost_equal(sum([v for k,v in distrib.items() if m in k]), 1, decimal=5, err_msg='"The concentration is not balanced"', verbose=True):
            pass
            print(sum([v for k,v in distrib.items() if m in k]))
    print(enzyme_concentration)
    print(distrib)

    # for each of the reactions within the dictionary, we need to set bounds for their fluxes
    #the lower bound will be 0, while the upper bound is determined by the concentration of the enzyme....
    #...which facilitates the reaction within a specific cell compartiment


    n = min(r_size,c_size)
    if n % 2 ==0:
        condition = n//2
    else:
        condition = n//2 + 1

    enzyme_concentration = {'k_AeAc': 1000, 'k_AcBc':2000}

    diffusion = {}

    for m, n in enzyme_concentration.items():
            for i in range(1, r_size+1):
                for j in range(1, c_size+1):
                    diffusion_id = m + str(i) + str(j)
                    for k, v in enzyme.items():
                        if diffusion_id == k:
                            if i == 1 or j == 1 or i == r_size or j == c_size:
                                diffusion[diffusion_id] = v * n
                            else:
                                for x in range(1, condition):
                                    if i != x and i != r_size - x + 1 and j != x and j != c_size - x + 1:
                                        diffusion[diffusion_id] = v * n * ((0.8) ** (x))

    print(diffusion)

    #some reactions have 2 indices and others have 3 (when the row numer has 2 digits)....
    #hence reactions id's can either have 7 or 8 characters
    #we are interested in comparing the last charachters from the reaction id with the last characters in enzyme id
    for k,v in react.items():
        for i, j in diffusion.items():
            if len(k) == 7:
                if k[-6:] == i[-6:]:
                    react[k].bounds=(0.0, j)
            else:
                if k[-7:] == i[-7:]:
                    react[k].bounds=(0.0, j)


    # let r/c be the row/column index of the cell part being deleted
    # if both the concentrations of the enzymes are being 0, then automatically the immediate neighbouring cell parts become external
    # notice that the deleted cell parts can only be external
    #this is just a suggestion for future work
    '''
    r = int(input("Deletion simulation row index "  + ":"))
    c = int(input("Deletion simulation column index "  + ":"))
    if r==0 or c==0:
        pass
    elif r!=1 and r!=r_size and c!=1 and c!=c_size:
        raise Exception("This cell part can not be deleted")
    else:
        for k, v in total_bounds.items():
            for i in range(r-1,r+1):
                for j in range(c-1,c+1):
                    if i == r and j == c:
                        n = str(r) + str(c)
                        if n in k:
                            react[k].bounds = (0.0, 0.0)
                    else:
                        m = str(i) + str(j)
                        if m in k:
                            react[k].bounds = (0.0, v)
    '''
    objective_functions = [v.flux_expression for k, v in react.items() if 'rBcBio' in k]
    quadratic_objective = test_model.problem.Objective(
        sum(objective_functions),
        direction='max')
    test_model.objective = quadratic_objective
    solution = test_model.optimize(objective_sense=None)
    print(solution)
    print({k:v.flux for k,v in react.items()})

    return solution, react

wb = load_workbook(filename = "C:\\Users\\Lenovo\\PycharmProjects\\systems_biology.xlsx")
wb_out = Workbook()
dest_filename = "C:\\Users\\Lenovo\\PycharmProjects\\results.xlsx"
sheet_name =[ "Model 1 Cell 1", "Model 1 Cell 2", "Model 1 Cell 3",
              "Model 2 Cell 1", "Model 2 Cell 2", "Model 2 Cell 3",
             "Model 3 Cell 1", "Model 3 Cell 2", "Model 3 Cell 3",
             "Model 4 Cell 1", "Model 4 Cell 2","Model 4 Cell 3",
             "Model 5 Cell 1", "Model 5 Cell 2","Model 5 Cell 3",
             "Model 6 Cell 1", "Model 6 Cell 2", "Model 6 Cell 3",
             "Model 7 Cell 1", "Model 7 Cell 2", "Model 7 Cell 3",
             "Model 8 Cell 1", "Model 8 Cell 2", "Model 8 Cell 3",
             "Model 9 Cell 1", "Model 9 Cell 2","Model 9 Cell 3",
             "Model 10 Cell 1", "Model 10 Cell 2","Model 10 Cell 3",
             "Model 11 Cell 1", "Model 11 Cell 2", "Model 11 Cell 3",
             "Model 12 Cell 1", "Model 12 Cell 2", "Model 12 Cell 3"]

results = {}
for i in sheet_name:
    sheet_ranges = wb[i]
    enzyme_dist = {}
    for row in range (6,36):
        enzyme_dist[sheet_ranges.cell(column=2,row=row).value] = sheet_ranges.cell(column=3,row=row).value
    for row in range (39,69):
        enzyme_dist[sheet_ranges.cell(column=2, row=row).value] = sheet_ranges.cell(column=3, row=row).value
    r_size = sheet_ranges.cell(column=3,row=2).value
    c_size = sheet_ranges.cell(column=3, row=3).value
    sol, fluxes = run_model(r_size, c_size, enzyme_dist)
    results[i] = [sol, fluxes]

    # Write to excel file
    ws = wb_out.create_sheet(title=i)
    row = 1
    for k,v in fluxes.items():
        ws['A' + str(row)] = k
        ws['B' + str(row)] = v.flux
        row += 1
    ws['A' + str(row)] = 'Solution'
    ws['B' + str(row)] = sol.objective_value
wb_out.save(filename = dest_filename)